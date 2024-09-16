# 初始化字典
FILE_PATH = "./stock.json"
import openpyxl
import json
def init():
    #建立一个数据库字典，用于存储每个月份各个城市的库存情况
    #城市有上海	广州	武汉	北京	济南	苏州	常熟	泰安	深圳	珠海
    #月份month有1-12月
    #库存商品有#品牌茶饮 品牌果汁 品牌咖啡  品牌汽水 品牌饮用水  网红茶饮 网红果汁  网红咖啡 网红汽水 网红饮用水 自营茶饮 自营咖啡 自营果汁 自营汽水 自营饮用水
    #开始
    citys = ['上海','广州','武汉','北京','济南','苏州','常熟','泰安','深圳','珠海']
    months = [1,2,3,4,5,6,7,8,9,10,11,12]
    goods = ['品牌茶饮','品牌果汁','品牌咖啡','品牌汽水','品牌饮用水','网红茶饮','网红果汁','网红咖啡','网红汽水','网红饮用水','自营茶饮','自营咖啡','自营果汁','自营汽水','自营饮用水']
    #建立一个字典，用于存储每个月份各个城市的库存情况
    stock = {}
    for city in citys:
        stock[city] = {}
        for month in months:
            stock[city][month] = {}
            for good in goods:
                stock[city][month][good] = 0
    #打印字典
    print(stock)
    #保存到stock.json文件，能以中文显示,没有则在当前目录下新建一个stock.json文件    
    import json
    with open(FILE_PATH,'w',encoding='utf-8') as f:
        json.dump(stock,f,ensure_ascii=False)
#增加库存
def add_stock(city,month,good,num):
    #读取stock.json文件
    import json
    with open(FILE_PATH,'r',encoding='utf-8') as f:
        stock = json.load(f)
    #增加库存
    stock[city][month][good] += num
    #保存到stock.json文件
    with open(FILE_PATH,'w',encoding='utf-8') as f:
        json.dump(stock,f,ensure_ascii=False)
    print('增加库存成功')
# 将当前的库存情况保存到所有城市输出到一个库存文件夹中,
# 每一个城市为单独的excel,文件名为该城市的名称.xlsx，
# 按月份分为12个sheet，每个sheet为一个月份的库存情况，每个sheet的行为城市，列为商品，单元格为库存数量
# 保存到stock文件夹中
def save_to_excel():
    #读取stock.json文件
    with open(FILE_PATH,'r',encoding='utf-8') as f:
        stock = json.load(f)
    #创建一个excel文件
    wb = openpyxl.Workbook()
    #删除默认的sheet
    wb.remove(wb.active)
    #遍历stock字典，将每个城市的库存情况保存到一个excel文件中
    for city in stock:
        #创建一个excel文件
        wb = openpyxl.Workbook()
        #删除默认的sheet
        wb.remove(wb.active)
        for month in stock[city]:
            #创建一个sheet
            ws = wb.create_sheet(month)
            #写入数据
            for i,good in enumerate(stock[city][month]):
                ws.cell(i+1,1,good)
                ws.cell(i+1,2,stock[city][month][good])
        #保存到stock文件夹中,没有要新建
        wb.save(f'./stock/{city}.xlsx')
    print('保存成功')
save_to_excel()





 
 
